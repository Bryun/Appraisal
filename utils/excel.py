from os import path
from re import search

from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.workbook import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from pandas import DataFrame


class Excel:
    def __init__(self, location: str, formula: bool = False):
        self.edit: bool = False
        self.location: str = location
        self.exists: bool = path.exists(location)
        self.alphabet: list = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R',
                               'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z']
        self.book = load_workbook(self.location, data_only=not formula) if self.exists else Workbook()

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        if self.edit:
            self.book.save(filename=self.location)

    @staticmethod
    def style(name: str = 'TableStyleMedium9'):
        return TableStyleInfo(name=name, showFirstColumn=False, showLastColumn=False, showRowStripes=True,
                              showColumnStripes=False)

    async def append(self, name: str, data: list[dict]):

        if data is None or len(data) == 0:
            print('No data to save.')
            return

        self.edit = True
        default: str = 'Sheet'

        if any(s == default for s in self.book.sheetnames):
            sheet = self.book.get_sheet_by_name(default)
            sheet.title = name
        elif any(s == name for s in self.book.sheetnames):
            sheet = self.book.get_sheet_by_name(name)
        else:
            sheet = self.book.create_sheet(title=name)
            sheet.title = name

        if sheet.tables is not None and len(sheet.tables) > 0:
            (identifier, range) = sheet.tables.items()[0]
            table = sheet.tables[identifier]
            number: int = int(search(r"\d+$", range).group(0))
            letter: str = search(r":([A-Z]+)\d+$", range).group(1)
            reference: str = f"A1:{letter}{len(data) + number}"

            table.ref = reference
            table.tableStyleInfo = self.style()

        else:
            sheet.append([k for k, v in data[0].items()])
            reference: str = f"A1:{self.alphabet[len(data[0].items()) - 1]}{len(data) + 1}"

            table = Table(displayName=name, ref=reference)
            table.tableStyleInfo = self.style()

            sheet.add_table(table)

        for row in data:
            sheet.append([v for k, v in row.items()])

    async def read(self, name: str = 'Sheet1', start: tuple = (), end: tuple = ()) -> DataFrame:
        self.edit = False
        sheet = self.book.get_sheet_by_name(name)

        headers: list = []
        rows: list = []

        for i in range(start[1], end[1]):
            row: list = []
            for letter in self.alphabet:

                if i == start[1]:
                    headers.append(sheet[f"{letter}{i}"].value)
                else:
                    row.append(sheet[f"{letter}{i}"].value)

                if letter == end[0]:
                    if i != start[1]:
                        rows.append(tuple(row))
                    break

        return DataFrame(rows, columns=headers)

    async def save_dataframe(self, name: str, data: DataFrame):
        self.edit = True

        if any(s == name for s in self.book.sheetnames):
            sheet = self.book.get_sheet_by_name(name)
        else:
            sheet = self.book.create_sheet(title=name)

        rows = dataframe_to_rows(data, index=False)

        for y, row in enumerate(rows, 1):
            for x, value in enumerate(row, 1):
                sheet.cell(row=y, column=x, value=value)

        reference: str = f"A1:{self.alphabet[len(data.columns) - 1]}{len(data.index) + 1}"
        table = Table(displayName=name, ref=reference)
        table.tableStyleInfo = self.style()

        sheet.add_table(table)

    def replicate_style(self, source: str, destination: str):
        pass

        # >
        #
        # def copy_style(src_cell, dest_cell):
        #     > dest_cell.font = src_cell.font
        #
        # > dest_cell.fill = src_cell.fill
        # > dest_cell.border = src_cell.border
        # > dest_cell.alignment = src_cell.alignment
        # > dest_cell.number_format = src_cell.number_format

    async def layout(self, start: tuple = (), end: tuple = ()) -> DataFrame:
        self.edit = False

        name: str = self.book.get_sheet_names()[0]
        sheet = self.book.get_sheet_by_name(name)
        headers: list = []
        rows: list = []

        for i in range(start[1], end[1]):
            row: list = []
            for letter in self.alphabet:

                if i == start[1]:
                    headers.append(sheet[f"{letter}{i}"].value)
                else:
                    row.append(sheet[f"{letter}{i}"].value)

                if letter == end[0]:
                    if i != start[1]:
                        rows.append(tuple(row))
                    break

        return DataFrame(rows, columns=headers)
